import os
from pathlib import Path
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.models import (
    User, StudentProfile, ContentItem, QuizAttempt,
    UserInteraction, Flashcard, ClassAssignment, SpacedRepetitionSchedule
)

# Output Excel path in the project root
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
DEFAULT_EXCEL_PATH = (PROJECT_ROOT / "edufeedia_database_records.xlsx").resolve()

def style_header_row(ws, title: str):
    """
    Applies styling to headers: dark blue fill, white bold text, centered alignment.
    """
    header_fill = PatternFill(start_color="1A2B4C", end_color="1A2B4C", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 24

def auto_fit_columns(ws):
    """
    Auto-adjusts column widths to fit content nicely.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 3))

def sync_database_to_excel(db: Session, export_path: str = None) -> str:
    """
    Reads all SQLite SQL tables and exports a comprehensive, structured, read-only Excel workbook (.xlsx).
    """
    target_path = Path(export_path) if export_path else DEFAULT_EXCEL_PATH
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Sheet: Users
    ws_users = wb.create_sheet(title="Users")
    ws_users.append(["User ID", "Full Name", "Email Address", "Role", "School Domain", "Is Verified", "Created Date"])
    users = db.query(User).all()
    for u in users:
        ws_users.append([
            u.id,
            f"{u.first_name} {u.last_name}",
            u.email,
            u.role.upper(),
            u.school.name if u.school else "Apex Academy",
            "YES" if u.is_verified else "NO",
            u.created_at.strftime("%Y-%m-%d %H:%M:%S") if u.created_at else ""
        ])
    style_header_row(ws_users, "Users")
    auto_fit_columns(ws_users)

    # 2. Sheet: Student_Profiles
    ws_students = wb.create_sheet(title="Student_Profiles")
    ws_students.append(["User ID", "Student Name", "Email", "Grade Level", "Section", "Board", "Total XP", "Streak (Days)", "Interests", "DOB"])
    profiles = db.query(StudentProfile).all()
    for sp in profiles:
        ws_students.append([
            sp.user_id,
            f"{sp.user.first_name} {sp.user.last_name}" if sp.user else "N/A",
            sp.user.email if sp.user else "N/A",
            sp.school_class.grade_level if sp.school_class else 10,
            sp.school_class.section_name if sp.school_class else "A",
            sp.board or "CBSE",
            sp.xp_score,
            sp.streak_count,
            ", ".join(sp.interests or []),
            str(sp.date_of_birth) if sp.date_of_birth else ""
        ])
    style_header_row(ws_students, "Student_Profiles")
    auto_fit_columns(ws_students)

    # 3. Sheet: Content_Catalog
    ws_content = wb.create_sheet(title="Content_Catalog")
    ws_content.append(["Content ID", "Module Title", "Subject", "Topic", "Grade", "Board", "Type", "Safety Score", "Edu Score", "Views", "Likes"])
    items = db.query(ContentItem).all()
    for c in items:
        ws_content.append([
            c.id,
            c.title,
            c.subject,
            c.topic,
            c.grade_level,
            c.board,
            c.type.upper(),
            c.safety_score,
            c.edu_score,
            c.view_count,
            c.like_count
        ])
    style_header_row(ws_content, "Content_Catalog")
    auto_fit_columns(ws_content)

    # 4. Sheet: Quiz_Attempts
    ws_quizzes = wb.create_sheet(title="Quiz_Attempts")
    ws_quizzes.append(["Attempt ID", "Student Name", "Student Email", "Score", "Max Score", "Accuracy %", "Graded Date"])
    attempts = db.query(QuizAttempt).order_by(QuizAttempt.completed_at.desc()).all()
    for qa in attempts:
        ws_quizzes.append([
            qa.id,
            f"{qa.student.first_name} {qa.student.last_name}" if qa.student else "Student",
            qa.student.email if qa.student else "N/A",
            qa.score,
            qa.max_score,
            f"{qa.accuracy_percentage:.1f}%",
            qa.completed_at.strftime("%Y-%m-%d %H:%M:%S") if qa.completed_at else ""
        ])
    style_header_row(ws_quizzes, "Quiz_Attempts")
    auto_fit_columns(ws_quizzes)

    # 5. Sheet: User_Interactions
    ws_inter = wb.create_sheet(title="User_Interactions")
    ws_inter.append(["Interaction ID", "User Email", "Content Title", "Interaction Signal", "Signal Weight", "Dwell Time (s)", "Timestamp"])
    interactions = db.query(UserInteraction).order_by(UserInteraction.created_at.desc()).all()
    for inter in interactions:
        ws_inter.append([
            inter.id,
            inter.user.email if inter.user else "User",
            inter.content_item.title if inter.content_item else "Lesson",
            inter.interaction_type.upper(),
            float(inter.weight),
            inter.dwell_time_seconds,
            inter.created_at.strftime("%Y-%m-%d %H:%M:%S") if inter.created_at else ""
        ])
    style_header_row(ws_inter, "User_Interactions")
    auto_fit_columns(ws_inter)

    # 6. Sheet: Flashcards
    ws_flash = wb.create_sheet(title="Flashcards")
    ws_flash.append(["Flashcard ID", "Subject", "Topic", "Front Formula / Cue", "Back Active Concept", "Recall Hint"])
    flashcards = db.query(Flashcard).all()
    for f in flashcards:
        ws_flash.append([
            f.id,
            f.subject,
            f.topic,
            f.front_text,
            f.back_text,
            f.hint or ""
        ])
    style_header_row(ws_flash, "Flashcards")
    auto_fit_columns(ws_flash)

    # 7. Sheet: Spaced_Schedules
    ws_spaced = wb.create_sheet(title="Spaced_Schedules")
    ws_spaced.append(["Schedule ID", "Student Email", "Subject", "Topic", "Interval (Days)", "Repetitions", "Easiness Factor", "Next Review Date"])
    schedules = db.query(SpacedRepetitionSchedule).all()
    for s in schedules:
        ws_spaced.append([
            s.id,
            s.student.email if s.student else "Student",
            s.subject,
            s.topic,
            s.interval_days,
            s.repetition_number,
            float(s.easiness_factor),
            str(s.next_review_date)
        ])
    style_header_row(ws_spaced, "Spaced_Schedules")
    auto_fit_columns(ws_spaced)

    # 8. Sheet: Class_Assignments
    ws_assign = wb.create_sheet(title="Class_Assignments")
    ws_assign.append(["Assignment ID", "Class Grade", "Teacher Email", "Title", "Instructions", "Due Date", "Created Date"])
    assignments = db.query(ClassAssignment).all()
    for a in assignments:
        ws_assign.append([
            a.id,
            f"Grade {a.school_class.grade_level}-{a.school_class.section_name}" if a.school_class else "Class",
            a.teacher.email if a.teacher else "Teacher",
            a.title,
            a.instructions or "",
            str(a.due_date) if a.due_date else "",
            a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else ""
        ])
    style_header_row(ws_assign, "Class_Assignments")
    auto_fit_columns(ws_assign)

    # Save workbook with safe Windows lock handling
    os.makedirs(target_path.parent, exist_ok=True)
    try:
        wb.save(target_path)
    except PermissionError:
        # File is locked because it is currently open in Microsoft Excel
        alt_path = target_path.with_name(f"{target_path.stem}_latest.xlsx")
        try:
            wb.save(alt_path)
            return str(alt_path)
        except Exception:
            pass
    return str(target_path)
